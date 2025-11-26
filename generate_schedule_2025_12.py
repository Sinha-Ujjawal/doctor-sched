from collections import defaultdict

import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side
from pandas import ExcelWriter

from doctor_schedule import all_shifts, generate_month_dates, generate_schedule, weeks
from doctors import (
    DR_ABHILASHA_MISHRA,
    DR_AMIT_TRIPATHI,
    DR_ASHISH_GUPTA,
    DR_ASHOK_KUMAR,
    DR_DEEPARK_SINGH,
    DR_HITESH_KUREEL,
    DR_KRITIKA_PRASAD,
    DR_MINAKSHI_MISHRA,
    DR_RAJAT_GUPTA,
    DR_RASHMI_SHARMA,
    DR_SAUMYA_SHUKLA,
    DR_SUNITA_OJHA,
    DR_VIKAS_VERMA,
)


def main():
    year = 2025
    month = 12
    all_doctors = [
        DR_MINAKSHI_MISHRA,
        DR_ABHILASHA_MISHRA,
        DR_AMIT_TRIPATHI,
        DR_RASHMI_SHARMA,
        DR_SAUMYA_SHUKLA,
        DR_KRITIKA_PRASAD,
        DR_ASHOK_KUMAR,
        DR_HITESH_KUREEL,
        DR_SUNITA_OJHA,
        DR_DEEPARK_SINGH,
        DR_RAJAT_GUPTA,
        DR_VIKAS_VERMA,
        DR_ASHISH_GUPTA,
    ]
    emos = [
        DR_MINAKSHI_MISHRA,
        DR_RASHMI_SHARMA,
        DR_KRITIKA_PRASAD,
        DR_HITESH_KUREEL,
        DR_SUNITA_OJHA,
        DR_DEEPARK_SINGH,
        DR_RAJAT_GUPTA,
        DR_VIKAS_VERMA,
        DR_ASHISH_GUPTA,
    ]
    dates = generate_month_dates(year, month)
    fixed_shifts = {}
    unavailable_shifts = defaultdict(list)
    leaves = []
    # Unavailable shifts
    for dt in dates:
        day = dt.day
        for emo in emos:
            unavailable_shifts[(emo, day)] = ["ot_duty"]
        if dt.weekday() == weeks.index("Mon"):
            unavailable_shifts[(DR_ABHILASHA_MISHRA, day)] = ["ot_duty"]
        if dt.weekday() == weeks.index("Sun"):
            unavailable_shifts[(DR_AMIT_TRIPATHI, day)] = ["morning", "evening"]
            unavailable_shifts[(DR_ABHILASHA_MISHRA, day)] = ["morning", "evening"]
            unavailable_shifts[(DR_SAUMYA_SHUKLA, day)] = ["morning", "evening"]
            unavailable_shifts[(DR_ASHOK_KUMAR, day)] = ["morning", "evening"]
            unavailable_shifts[(DR_DEEPARK_SINGH, day)] = all_shifts
        if dt.weekday() == weeks.index("Wed"):
            unavailable_shifts[(DR_ABHILASHA_MISHRA, day)] = ["night"]

    """
    Dr Rashmi 10-20
    Dr Kritika taking full month EL
    Dr Amit 23-27
    Dr Minakshi 3-8
    Dr Rajat Gupta 30th Dec to 1st Jan
    Dr Ashish is on leave 8,9
    """
    unofficial_leaves = [
        [DR_RASHMI_SHARMA, range(10, 20 + 1)],
        [DR_KRITIKA_PRASAD, range(1, 31 + 1)],
        [DR_AMIT_TRIPATHI, range(23, 27 + 1)],
        [DR_MINAKSHI_MISHRA, range(3, 8 + 1)],
        [DR_RAJAT_GUPTA, [30, 31]],
        [DR_ASHISH_GUPTA, [8, 9]],
    ]

    for doctor, days in [*leaves, *unofficial_leaves]:
        for day in days:
            unavailable_shifts[(doctor, day)] = all_shifts
    sunday_doctor_rotation_order = [
        DR_ABHILASHA_MISHRA,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_ASHOK_KUMAR,
    ]
    sunday_doctor_rotation_start = 2
    for dt in dates:
        day = dt.day
        week = dt.weekday()
        if week == weeks.index("Sun"):
            fixed_shifts[
                (sunday_doctor_rotation_order[sunday_doctor_rotation_start], day)
            ] = ["ot_duty"]
            sunday_doctor_rotation_start = (sunday_doctor_rotation_start + 1) % len(
                sunday_doctor_rotation_order
            )
        if week in (weeks.index("Mon"),):
            fixed_shifts[(DR_ASHOK_KUMAR, day)] = ["ot_duty"]
        if week in (weeks.index("Tue"),):
            fixed_shifts[(DR_AMIT_TRIPATHI, day)] = ["ot_duty"]
        if week in (weeks.index("Thu"),):
            fixed_shifts[(DR_ABHILASHA_MISHRA, day)] = ["ot_duty", "night"]
        if week in (weeks.index("Fri"),):
            fixed_shifts[(DR_SAUMYA_SHUKLA, day)] = ["ot_duty"]

    def my_custom_constraints(model, shift_vars):
        # DR_RASHMI_SHARMA will do evening duty on 1, 4, 7, ...
        #                       and night duty on  2, 5, 8, ...
        # And remember she isn't available from 10-20
        DR_RASHMI_SHARMA_INDEX = all_doctors.index(DR_RASHMI_SHARMA)
        for dt in dates:
            if 10 <= dt.day <= 20:
                continue
            if dt.weekday() in (weeks.index("Sun"), weeks.index("Sat")):
                continue
            if (dt.day - 1) % 3 == 0:  # 1, 4, 7...
                model.Add(shift_vars[(DR_RASHMI_SHARMA_INDEX, dt.day, "evening")] == 1)
            elif (
                dt.day <= 20
                and dt.day >= 2
                and dt.day not in (5, 6, 7)
                and (dt.day - 2) % 3 == 0
            ):  # 2, 5, 8...
                model.Add(shift_vars[(DR_RASHMI_SHARMA_INDEX, dt.day, "night")] == 1)
            elif dt.day <= 20:
                model.Add(
                    sum(
                        shift_vars[(DR_RASHMI_SHARMA_INDEX, dt.day, shift)]
                        for shift in all_shifts
                    )
                    <= 0
                )

        # DR_ASHOK_KUMAR will do ot-duty with night on same day
        DR_ASHOK_KUMAR_INDEX = all_doctors.index(DR_ASHOK_KUMAR)
        for dt in dates:
            week = dt.weekday()
            day = dt.day
            if week == weeks.index("Mon"):
                model.Add(
                    shift_vars[(DR_ASHOK_KUMAR_INDEX, dt.day, "night")] == 1
                ).OnlyEnforceIf(shift_vars[(DR_ASHOK_KUMAR_INDEX, dt.day, "ot_duty")])
                model.Add(
                    shift_vars[(DR_ASHOK_KUMAR_INDEX, dt.day, "morning")] == 0
                ).OnlyEnforceIf(shift_vars[(DR_ASHOK_KUMAR_INDEX, dt.day, "ot_duty")])
                model.Add(
                    shift_vars[(DR_ASHOK_KUMAR_INDEX, dt.day, "evening")] == 0
                ).OnlyEnforceIf(shift_vars[(DR_ASHOK_KUMAR_INDEX, dt.day, "ot_duty")])
            elif week != weeks.index("Sun"):
                model.Add(
                    sum(
                        shift_vars[(DR_ASHOK_KUMAR_INDEX, dt.day, shift)]
                        for shift in all_shifts
                    )
                    <= 1
                )

        # EMO Doctor who is doing Friday night is supposed to do morning and evening on coming Sunday.
        for doc_idx, doc in enumerate(all_doctors):
            if doc not in emos:
                continue
            for dt in dates:
                week = dt.weekday()
                day = dt.day
                day_plus_2 = day + 2
                if week == weeks.index("Fri") and day_plus_2 <= len(dates):
                    model.Add(
                        shift_vars[(doc_idx, day_plus_2, "morning")] == 1
                    ).OnlyEnforceIf(shift_vars[(doc_idx, day, "night")])
                    model.Add(
                        shift_vars[(doc_idx, day_plus_2, "evening")] == 1
                    ).OnlyEnforceIf(shift_vars[(doc_idx, day, "night")])

        # DR_ABHILASHA_MISHRA will do ot-duty with night on same day
        DR_ABHILASHA_MISHRA_INDEX = all_doctors.index(DR_ABHILASHA_MISHRA)
        # model.Add(sum(
        #     shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, "night")]
        #     for dt in dates
        #     if dt.weekday() == weeks.index("Thu")
        # ) >= 1)
        for dt in dates:
            week = dt.weekday()
            day = dt.day
            if week in (weeks.index("Thu"), weeks.index("Sun")):
                model.Add(
                    shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, "night")] == 1
                ).OnlyEnforceIf(
                    shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, "ot_duty")]
                )
                model.Add(
                    shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, "morning")] == 0
                ).OnlyEnforceIf(
                    shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, "ot_duty")]
                )
                model.Add(
                    shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, "evening")] == 0
                ).OnlyEnforceIf(
                    shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, "ot_duty")]
                )
            elif week != weeks.index("Sun"):
                model.Add(
                    sum(
                        shift_vars[(DR_ABHILASHA_MISHRA_INDEX, dt.day, shift)]
                        for shift in all_shifts
                    )
                    <= 1
                )

    max_night_shifts = {
        # EMOs
        DR_SUNITA_OJHA: 0,
        DR_KRITIKA_PRASAD: 0,
        DR_MINAKSHI_MISHRA: 2,
        DR_HITESH_KUREEL: 2,
        DR_DEEPARK_SINGH: 3,
        DR_RAJAT_GUPTA: 3,
        DR_VIKAS_VERMA: 3,
        DR_ASHISH_GUPTA: 3,
        # Surgeons
        DR_ABHILASHA_MISHRA: 5,
        DR_SAUMYA_SHUKLA: 2,
        DR_AMIT_TRIPATHI: 2,
        DR_ASHOK_KUMAR: 5,
    }
    max_evening_shifts = {}
    for doc in all_doctors:
        if doc not in emos or doc == DR_SUNITA_OJHA:
            max_evening_shifts[doc] = 0
        elif doc == DR_RASHMI_SHARMA:
            max_evening_shifts[doc] = 6
        elif doc == DR_MINAKSHI_MISHRA:
            max_evening_shifts[doc] = 7
        else:
            max_evening_shifts[doc] = 4
    max_morning_shifts = {}
    rough_morning_upper_limit = (31 // (len(emos) - 2)) + 1
    for doc in all_doctors:
        if doc not in emos:
            max_morning_shifts[doc] = 0
        elif doc == DR_SUNITA_OJHA:
            max_morning_shifts[doc] = 0
        elif doc == DR_RASHMI_SHARMA:
            max_morning_shifts[doc] = 1
        else:
            max_morning_shifts[doc] = rough_morning_upper_limit
    minmax_ot_duty_shifts = {
        DR_ASHOK_KUMAR: (7, 8),
        DR_ABHILASHA_MISHRA: (7, 8),
        DR_AMIT_TRIPATHI: (7, 9),
        DR_SAUMYA_SHUKLA: (7, 8),
    }
    avoid_shift_collision = []
    for dt in dates:
        day = dt.day
        week = dt.weekday()
        avoid_shift_collision.extend(
            [
                (
                    DR_SAUMYA_SHUKLA,
                    "ot_duty",
                    day,
                    DR_ABHILASHA_MISHRA,
                    "morning",
                    day,
                ),
                (
                    DR_SAUMYA_SHUKLA,
                    "ot_duty",
                    day,
                    DR_ABHILASHA_MISHRA,
                    "evening",
                    day,
                ),
                (
                    DR_SAUMYA_SHUKLA,
                    "ot_duty",
                    day,
                    DR_ABHILASHA_MISHRA,
                    "night",
                    day,
                ),
                (
                    DR_ABHILASHA_MISHRA,
                    "ot_duty",
                    day,
                    DR_SAUMYA_SHUKLA,
                    "morning",
                    day,
                ),
                (
                    DR_ABHILASHA_MISHRA,
                    "ot_duty",
                    day,
                    DR_SAUMYA_SHUKLA,
                    "evening",
                    day,
                ),
                (
                    DR_ABHILASHA_MISHRA,
                    "ot_duty",
                    day,
                    DR_SAUMYA_SHUKLA,
                    "night",
                    day,
                ),
            ]
        )
    first_night_off = DR_AMIT_TRIPATHI
    print("Generating schedule...")
    solution_maybe = generate_schedule(
        doctors=all_doctors,
        year=year,
        month=month,
        fixed_shifts=fixed_shifts,
        unavailable_shifts=unavailable_shifts,
        first_night_off=first_night_off,
        max_night_shifts=max_night_shifts,
        max_evening_shifts=max_evening_shifts,
        max_morning_shifts=max_morning_shifts,
        minmax_ot_duty_shifts=minmax_ot_duty_shifts,
        wed_ot_duty_rotation_size=4,
        sat_ot_duty_rotation_size=None,
        sun_ot_duty_rotation_size=4,
        same_sat_and_sun_ot_duty=True,
        sun_morning_evening_duty_rotation_size=4,
        avoid_shift_collision=avoid_shift_collision,
        custom_constraints=my_custom_constraints,
        doctors_who_wants_do_more_shifts_per_day=[DR_ABHILASHA_MISHRA, DR_ASHOK_KUMAR],
    )
    if solution_maybe is not None:
        df_schedule, df_stats = solution_maybe
        excel_output = f"schedule_{year}_{month}.xlsx"
        df_schedule.columns = [
            "Date",
            "Days",
            "OT Duty",
            "Morning",
            "Evening",
            "Night",
            "Night Off",
        ]
        df_schedule["Date"] = df_schedule["Date"].apply(
            lambda dt: dt.strftime("%d/%m/%Y")
        )
        print(f"One solution found!, writing it to {excel_output}")

        # Calculate statistics from df_schedule
        doctor_ot_counts = {}
        doctor_night_counts = {}

        for _, row in df_schedule.iterrows():
            # Count OT duties
            if pd.notna(row["OT Duty"]) and row["OT Duty"]:  # type: ignore
                doctors_in_ot = [d.strip() for d in str(row["OT Duty"]).split(",")]
                for doc in doctors_in_ot:
                    if doc:
                        doctor_ot_counts[doc] = doctor_ot_counts.get(doc, 0) + 1

            # Count nights
            if pd.notna(row["Night"]) and row["Night"]:  # type: ignore
                doctors_in_night = [d.strip() for d in str(row["Night"]).split(",")]
                for doc in doctors_in_night:
                    if doc:
                        doctor_night_counts[doc] = doctor_night_counts.get(doc, 0) + 1

        with ExcelWriter(excel_output, engine="openpyxl") as xlw:  # type: ignore
            df_schedule.to_excel(index=False, excel_writer=xlw, sheet_name="schedule")
            df_stats.to_excel(index=False, excel_writer=xlw, sheet_name="stats")

            # Format the schedule sheet
            ws = xlw.sheets["schedule"]

            # Remove gridlines
            ws.sheet_view.showGridLines = False

            # Color header row
            header_fill = PatternFill(
                start_color="000000", end_color="000000", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            # Create thin border
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border

            # Color rows based on day of week
            green_fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid"
            )
            yellow_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            orange_fill = PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                day_name = row[1].value
                if day_name == "Wed":
                    for cell in row:
                        cell.fill = green_fill
                        cell.border = thin_border
                elif day_name == "Sat":
                    for cell in row:
                        cell.fill = yellow_fill
                        cell.border = thin_border
                elif day_name == "Sun":
                    for cell in row:
                        cell.fill = orange_fill
                        cell.border = thin_border
                else:
                    for cell in row:
                        cell.border = thin_border

            # Auto-fit column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as _:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add statistics below the schedule
            stats_start_row = ws.max_row + 3

            # Doctor statistics header
            ws.cell(row=stats_start_row, column=1, value="Doctor").fill = header_fill
            ws.cell(row=stats_start_row, column=1).font = header_font
            ws.cell(row=stats_start_row, column=1).border = thin_border

            ws.cell(
                row=stats_start_row, column=2, value="No. of OTs"
            ).fill = header_fill
            ws.cell(row=stats_start_row, column=2).font = header_font
            ws.cell(row=stats_start_row, column=2).border = thin_border

            ws.cell(
                row=stats_start_row, column=3, value="No. of Nights"
            ).fill = header_fill
            ws.cell(row=stats_start_row, column=3).font = header_font
            ws.cell(row=stats_start_row, column=3).border = thin_border

            # Leaves header
            ws.cell(row=stats_start_row, column=5, value="Leaves").fill = header_fill
            ws.cell(row=stats_start_row, column=5).font = header_font
            ws.cell(row=stats_start_row, column=5).border = thin_border

            # Add doctor statistics
            current_row = stats_start_row + 1
            total_ots = 0
            total_nights = 0

            # Get all unique doctors from the calculated stats
            all_doc_names = set(doctor_ot_counts.keys()) | set(
                doctor_night_counts.keys()
            )

            for doctor in sorted(all_doc_names):
                ot_count = doctor_ot_counts.get(doctor, 0)
                night_count = doctor_night_counts.get(doctor, 0)

                ws.cell(row=current_row, column=1, value=doctor).border = thin_border
                if ot_count > 0:
                    ws.cell(
                        row=current_row, column=2, value=ot_count
                    ).border = thin_border
                    total_ots += ot_count
                else:
                    ws.cell(row=current_row, column=2, value="-").border = thin_border

                if night_count > 0:
                    ws.cell(
                        row=current_row, column=3, value=night_count
                    ).border = thin_border
                    total_nights += night_count
                else:
                    ws.cell(row=current_row, column=3, value="-").border = thin_border

                current_row += 1

            # Add total row
            ws.cell(row=current_row, column=1, value="Total").fill = header_fill
            ws.cell(row=current_row, column=1).font = header_font
            ws.cell(row=current_row, column=1).border = thin_border

            ws.cell(row=current_row, column=2, value=total_ots).fill = header_fill
            ws.cell(row=current_row, column=2).font = header_font
            ws.cell(row=current_row, column=2).border = thin_border

            ws.cell(row=current_row, column=3, value=total_nights).fill = header_fill
            ws.cell(row=current_row, column=3).font = header_font
            ws.cell(row=current_row, column=3).border = thin_border

            # Add leaves information from the leaves object
            leaves_row = stats_start_row + 1
            for doctor, leave_days in leaves:
                doctor_name = str(doctor)  # Convert doctor object to string
                leave_days_str = ", ".join(map(str, leave_days))

                ws.cell(
                    row=leaves_row, column=5, value=doctor_name
                ).border = thin_border
                ws.cell(
                    row=leaves_row, column=6, value=leave_days_str
                ).border = thin_border
                leaves_row += 1
    else:
        print("No solutions found. Try adjusting the constraints")


if __name__ == "__main__":
    main()
