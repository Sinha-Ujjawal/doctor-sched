from collections import defaultdict

import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side
from pandas import ExcelWriter

from doctor_schedule import (
    Doctor,
    all_shifts,
    generate_month_dates,
    generate_schedule,
    weeks,
)
from doctors import (
    DR_AKASHDEEP_GUPTA,
    DR_AMIT_TRIPATHI,
    DR_HITESH_KUREEL,
    DR_KRITIKA_PRASAD,
    DR_MADHURI_TRIPATHI,
    DR_MINAKSHI_MISHRA,
    DR_RAJAT_GUPTA,
    DR_RASHMI_SHARMA,
    DR_SAUMYA_SHUKLA,
    DR_VIKAS_VERMA,
)


def main():
    """
    NOTES:
        DR_SAUMYA_SHUKLA is on leave on 4 and 5
        DR_KRITIKA_PRASAD is on leave entire September
        DR_AMIT_TRIPATHI is on leave on 23, 24 and 25
        DR_MADHURI_TRIPATHI is on leave on 23, 24 and 25
        DR_RASHMI_SHARMA is on leave between 4-8
    """

    year = 2026
    month = 9
    all_doctors = [
        DR_MINAKSHI_MISHRA,
        DR_AMIT_TRIPATHI,
        DR_RASHMI_SHARMA,
        DR_SAUMYA_SHUKLA,
        DR_KRITIKA_PRASAD,
        DR_HITESH_KUREEL,
        DR_AKASHDEEP_GUPTA,
        DR_RAJAT_GUPTA,
        DR_VIKAS_VERMA,
        DR_MADHURI_TRIPATHI,
    ]
    emos = [
        DR_MINAKSHI_MISHRA,
        DR_RASHMI_SHARMA,
        DR_KRITIKA_PRASAD,
        DR_HITESH_KUREEL,
        DR_AKASHDEEP_GUPTA,
        DR_RAJAT_GUPTA,
        DR_VIKAS_VERMA,
        DR_MADHURI_TRIPATHI,
    ]
    dates = generate_month_dates(year, month)
    fixed_shifts = {}
    unavailable_shifts = defaultdict(list)
    leaves = []
    # Unavailable shifts
    for dt in dates:
        day = dt.day
        for emo in emos:
            unavailable_shifts[(emo, day)] = ["ot_duty"]

        if dt.weekday() == weeks.index("Sun"):
            unavailable_shifts[(DR_AMIT_TRIPATHI, day)] = ["morning", "evening"]
            unavailable_shifts[(DR_SAUMYA_SHUKLA, day)] = ["morning", "evening"]
        else:
            unavailable_shifts[(DR_MADHURI_TRIPATHI, day)] = ["morning"]

    def range_list(lo: int, hi: int) -> list[int]:
        return list(range(lo, hi + 1))

    unofficial_leaves: list[tuple[str, list[int]]] = [
        (DR_SAUMYA_SHUKLA, [4, 5]),
        (DR_KRITIKA_PRASAD, range_list(1, 30)),
        (DR_AMIT_TRIPATHI, [23, 24, 25]),
        (DR_MADHURI_TRIPATHI, [23, 24, 25]),
        (DR_RASHMI_SHARMA, range_list(4, 8)),
    ]
    for doctor, days in [*leaves, *unofficial_leaves]:
        for day in days:
            unavailable_shifts[(doctor, day)] = all_shifts

    sunday_morning_and_evening_shifts = [
        # 6, 13, 20, 27
        DR_HITESH_KUREEL,
        DR_VIKAS_VERMA,
        DR_AKASHDEEP_GUPTA,
        DR_MADHURI_TRIPATHI,
    ]
    sunday_night_shifts = [
        # 6, 13, 20, 27
        DR_VIKAS_VERMA,
        DR_AKASHDEEP_GUPTA,
        DR_MADHURI_TRIPATHI,
        DR_RASHMI_SHARMA,
    ]
    sundays = [dt.day for dt in dates if dt.weekday() == weeks.index("Sun")]
    assert len(sunday_morning_and_evening_shifts) == len(sundays)
    assert len(sunday_night_shifts) == len(sundays)
    for day, morning_evening_doc, night_doc in zip(
        sundays, sunday_morning_and_evening_shifts, sunday_night_shifts
    ):
        fixed_shifts[(morning_evening_doc, day)] = ["morning", "evening"]
        fixed_shifts[(night_doc, day)] = ["night"]

    ot_duty_fixed_list = [
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_AMIT_TRIPATHI,  # DR_SAUMYA_SHUKLA, -- As DR_SAUMYA_SHUKLA is on leave
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_SAUMYA_SHUKLA,  # DR_AMIT_TRIPATHI, -- As DR_AMIT_TRIPATHI is on leave
        DR_SAUMYA_SHUKLA,
        DR_SAUMYA_SHUKLA,  # DR_AMIT_TRIPATHI, -- As DR_AMIT_TRIPATHI is on leave
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        DR_AMIT_TRIPATHI,
        DR_SAUMYA_SHUKLA,
        # DR_AMIT_TRIPATHI,
    ]

    # assert len(ot_duty_fixed_list) == 31
    assert len(ot_duty_fixed_list) == 30

    for day, doc in enumerate(ot_duty_fixed_list, 1):
        fixed_shifts[(doc, day)] = ["ot_duty"]

    def my_custom_constraints(model, shift_vars):
        # When DR_MADHURI_TRIPATHI does night, then Morning will be either by DR_RASHMI_SHARMA and Evening will be by DR_MINAKSHI_MISHRA or vice-versa
        DR_MADHURI_TRIPATHI_INDEX = all_doctors.index(DR_MADHURI_TRIPATHI)
        DR_RASHMI_SHARMA_INDEX = all_doctors.index(DR_RASHMI_SHARMA)
        DR_MINAKSHI_MISHRA_INDEX = all_doctors.index(DR_MINAKSHI_MISHRA)
        for dt in dates:
            if dt.weekday() != weeks.index("Sun"):
                model.Add(
                    (
                        shift_vars[(DR_RASHMI_SHARMA_INDEX, dt.day, "evening")]
                        + shift_vars[(DR_MINAKSHI_MISHRA_INDEX, dt.day, "evening")]
                    )
                    >= 1
                ).OnlyEnforceIf(
                    shift_vars[(DR_MADHURI_TRIPATHI_INDEX, dt.day, "night")]
                )

    minmax_night_shifts = {
        # EMOs
        DR_KRITIKA_PRASAD: (0, 0),
        DR_HITESH_KUREEL: (3, 5),
        DR_RAJAT_GUPTA: (3, 5),
        DR_AKASHDEEP_GUPTA: (3, 4),
        DR_VIKAS_VERMA: (3, 4),
        DR_MADHURI_TRIPATHI: (4, 5),
        # Senior EMOs
        DR_RASHMI_SHARMA: (3, 4),
        DR_MINAKSHI_MISHRA: (3, 4),
        # Surgeons
        DR_SAUMYA_SHUKLA: (0, 0),
        DR_AMIT_TRIPATHI: (0, 0),
    }
    minmax_evening_shifts: dict[Doctor, tuple[int, int]] = {}
    for doc in all_doctors:
        if doc == DR_MADHURI_TRIPATHI:
            minmax_evening_shifts[doc] = (20, 31)
        elif doc in [DR_MINAKSHI_MISHRA]:
            minmax_evening_shifts[doc] = (0, 3)
        elif doc in [DR_RASHMI_SHARMA]:
            minmax_evening_shifts[doc] = (0, 3)
        elif doc == DR_KRITIKA_PRASAD:
            minmax_evening_shifts[doc] = (0, 0)
        elif doc in emos:
            minmax_evening_shifts[doc] = (0, 1)
        else:
            minmax_evening_shifts[doc] = (0, 0)
    minmax_morning_shifts = {}
    for doc in all_doctors:
        if doc == DR_MADHURI_TRIPATHI:
            minmax_morning_shifts[doc] = (1, 1)
        elif doc == DR_KRITIKA_PRASAD:
            minmax_morning_shifts[doc] = (0, 0)
        elif doc in emos:
            minmax_morning_shifts[doc] = (4, 5)
        else:
            minmax_morning_shifts[doc] = (0, 0)
    minmax_ot_duty_shifts = {
        # DR_AMIT_TRIPATHI: (4, 15),
        # DR_SAUMYA_SHUKLA: (4, 15),
    }
    avoid_shift_collision = []
    for dt in dates:
        day = dt.day
        # week = dt.weekday()
        pass
        # if week == weeks.index("Sun"):
        #     avoid_shift_collision.extend(
        #         [
        #             (
        #                 DR_AMIT_TRIPATHI,
        #                 "ot_duty",
        #                 day,
        #                 DR_MADHURI_TRIPATHI,
        #                 "morning",
        #                 day,
        #             ),
        #             (
        #                 DR_AMIT_TRIPATHI,
        #                 "ot_duty",
        #                 day,
        #                 DR_MADHURI_TRIPATHI,
        #                 "evening",
        #                 day,
        #             ),
        #         ]
        #     )
    first_night_off = DR_MINAKSHI_MISHRA
    print("Generating schedule...")
    solution_maybe = generate_schedule(
        doctors=all_doctors,
        year=year,
        month=month,
        fixed_shifts=fixed_shifts,
        unavailable_shifts=unavailable_shifts,
        first_night_off=first_night_off,
        minmax_night_shifts=minmax_night_shifts,
        minmax_evening_shifts=minmax_evening_shifts,
        minmax_morning_shifts=minmax_morning_shifts,
        minmax_ot_duty_shifts=minmax_ot_duty_shifts,
        wed_ot_duty_rotation_size=None,
        sat_ot_duty_rotation_size=None,
        sun_ot_duty_rotation_size=None,
        same_sat_and_sun_ot_duty=False,
        sun_morning_evening_duty_rotation_size=None,
        avoid_shift_collision=avoid_shift_collision,
        custom_constraints=my_custom_constraints,
        sun_same_doctor_ot_and_night=False,
    )
    if solution_maybe is not None:
        df_schedule, df_stats = solution_maybe
        excel_output = f"schedule_{year}_{month}.xlsx"
        df_schedule.columns = [
            "Date",
            "Days",
            "OT Duty",
            "Morning",
            "Evening",
            "Night",
            "Night Off",
        ]
        df_schedule["Date"] = df_schedule["Date"].apply(
            lambda dt: dt.strftime("%d/%m/%Y")
        )
        print(f"One solution found!, writing it to {excel_output}")

        # Calculate statistics from df_schedule
        doctor_ot_counts = {}
        doctor_night_counts = {}

        for _, row in df_schedule.iterrows():
            # Count OT duties
            if pd.notna(row["OT Duty"]) and row["OT Duty"]:  # type: ignore
                doctors_in_ot = [d.strip() for d in str(row["OT Duty"]).split(",")]
                for doc in doctors_in_ot:
                    if doc:
                        doctor_ot_counts[doc] = doctor_ot_counts.get(doc, 0) + 1

            # Count nights
            if pd.notna(row["Night"]) and row["Night"]:  # type: ignore
                doctors_in_night = [d.strip() for d in str(row["Night"]).split(",")]
                for doc in doctors_in_night:
                    if doc:
                        doctor_night_counts[doc] = doctor_night_counts.get(doc, 0) + 1

        with ExcelWriter(excel_output, engine="openpyxl") as xlw:  # type: ignore
            df_schedule.to_excel(index=False, excel_writer=xlw, sheet_name="schedule")
            df_stats.to_excel(index=False, excel_writer=xlw, sheet_name="stats")

            def _populate_unavailable_leaves():
                unofficial_leaves_rows = []
                for doctor, leaves in unofficial_leaves:
                    current_low = None
                    current_high = None
                    ranges = []
                    for leave in leaves:
                        if current_low is None or current_high is None:
                            current_low = current_high = leave
                        elif leave == current_high + 1:
                            current_high = leave
                        else:
                            ranges.append((current_low, current_high))
                            current_low = current_high = leave
                    if current_low is not None and current_high is not None:
                        ranges.append((current_low, current_high))
                    unofficial_leaves_rows.append(
                        [
                            doctor,
                            ", ".join(
                                str(lb) if lb == ub else f"{lb}-{ub}"
                                for lb, ub in ranges
                            ),
                        ]
                    )
                pd.DataFrame(
                    unofficial_leaves_rows,
                    columns=["Doctor", "Unavailability"],  # type: ignore
                ).to_excel(
                    index=False, excel_writer=xlw, sheet_name="unofficial_leaves"
                )

            _populate_unavailable_leaves()

            # Format the schedule sheet
            ws = xlw.sheets["schedule"]

            # Remove gridlines
            ws.sheet_view.showGridLines = False

            # Color header row
            header_fill = PatternFill(
                start_color="000000", end_color="000000", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            # Create thin border
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border

            # Color rows based on day of week
            # green_fill = PatternFill(
            #     start_color="92D050", end_color="92D050", fill_type="solid"
            # )
            # yellow_fill = PatternFill(
            #     start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            # )
            orange_fill = PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                day_name = row[1].value
                # if day_name == "Wed":
                #     for cell in row:
                #         cell.fill = green_fill
                #         cell.border = thin_border
                # elif day_name == "Sat":
                #     for cell in row:
                #         cell.fill = yellow_fill
                #         cell.border = thin_border
                if day_name == "Sun":
                    for cell in row:
                        cell.fill = orange_fill
                        cell.border = thin_border
                else:
                    for cell in row:
                        cell.border = thin_border

            # Auto-fit column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as _:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add statistics below the schedule
            stats_start_row = ws.max_row + 3

            # Doctor statistics header
            ws.cell(row=stats_start_row, column=1, value="Doctor").fill = header_fill
            ws.cell(row=stats_start_row, column=1).font = header_font
            ws.cell(row=stats_start_row, column=1).border = thin_border

            ws.cell(
                row=stats_start_row, column=2, value="No. of OTs"
            ).fill = header_fill
            ws.cell(row=stats_start_row, column=2).font = header_font
            ws.cell(row=stats_start_row, column=2).border = thin_border

            ws.cell(
                row=stats_start_row, column=3, value="No. of Nights"
            ).fill = header_fill
            ws.cell(row=stats_start_row, column=3).font = header_font
            ws.cell(row=stats_start_row, column=3).border = thin_border

            # Leaves header
            ws.cell(row=stats_start_row, column=5, value="Leaves").fill = header_fill
            ws.cell(row=stats_start_row, column=5).font = header_font
            ws.cell(row=stats_start_row, column=5).border = thin_border

            # Add doctor statistics
            current_row = stats_start_row + 1
            total_ots = 0
            total_nights = 0

            # Get all unique doctors from the calculated stats
            all_doc_names = set(doctor_ot_counts.keys()) | set(
                doctor_night_counts.keys()
            )

            for doctor in sorted(all_doc_names):
                ot_count = doctor_ot_counts.get(doctor, 0)
                night_count = doctor_night_counts.get(doctor, 0)

                ws.cell(row=current_row, column=1, value=doctor).border = thin_border
                if ot_count > 0:
                    ws.cell(
                        row=current_row, column=2, value=ot_count
                    ).border = thin_border
                    total_ots += ot_count
                else:
                    ws.cell(row=current_row, column=2, value="-").border = thin_border

                if night_count > 0:
                    ws.cell(
                        row=current_row, column=3, value=night_count
                    ).border = thin_border
                    total_nights += night_count
                else:
                    ws.cell(row=current_row, column=3, value="-").border = thin_border

                current_row += 1

            # Add total row
            ws.cell(row=current_row, column=1, value="Total").fill = header_fill
            ws.cell(row=current_row, column=1).font = header_font
            ws.cell(row=current_row, column=1).border = thin_border

            ws.cell(row=current_row, column=2, value=total_ots).fill = header_fill
            ws.cell(row=current_row, column=2).font = header_font
            ws.cell(row=current_row, column=2).border = thin_border

            ws.cell(row=current_row, column=3, value=total_nights).fill = header_fill
            ws.cell(row=current_row, column=3).font = header_font
            ws.cell(row=current_row, column=3).border = thin_border

            # Add leaves information from the leaves object
            leaves_row = stats_start_row + 1
            for doctor, leave_days in leaves:
                doctor_name = str(doctor)  # Convert doctor object to string
                leave_days_str = ", ".join(map(str, leave_days))

                ws.cell(
                    row=leaves_row, column=5, value=doctor_name
                ).border = thin_border
                ws.cell(
                    row=leaves_row, column=6, value=leave_days_str
                ).border = thin_border
                leaves_row += 1
    else:
        print("No solutions found. Try adjusting the constraints")


if __name__ == "__main__":
    main()
